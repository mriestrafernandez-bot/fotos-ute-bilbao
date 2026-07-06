import base64
import io
import json
import os
import re
from copy import copy
from datetime import date
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Font
import pandas as pd
import streamlit as st
import xlrd
import xlwt
from xlutils.copy import copy as copy_xls_workbook
from dotenv import load_dotenv
from openai import OpenAI
from PIL import Image, ImageOps


load_dotenv()

APP_TITLE = "Fotos UTE Bilbao"
DEFAULT_MODEL = os.getenv("OPENAI_MODEL", "gpt-5-mini")
DOUBTFUL_CONFIDENCE = 0.8
DIRECT_CHUNK_SIZE = 6


def image_to_data_url(image: Image.Image) -> str:
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return f"data:image/png;base64,{encoded}"


def image_content(image: Image.Image) -> dict:
    return {"type": "input_image", "image_url": image_to_data_url(image), "detail": "high"}


def extract_json(text: str) -> dict:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned).strip()
        cleaned = re.sub(r"```$", "", cleaned).strip()

    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
        if not match:
            raise
        return json.loads(match.group(0))


def read_image_with_openai(
    client: OpenAI,
    image: Image.Image,
    model: str,
    month_hint: str,
) -> list[dict]:
    prompt = f"""
Eres un extractor de datos para una hoja mensual.
Lee esta fotografia JPEG. Contiene fechas y numeros escritos a mano junto a cada fecha.

Instrucciones:
- Extrae una fila por cada fecha visible que tenga un numero manuscrito asociado.
- Devuelve solo JSON valido, sin markdown.
- Si una fecha no incluye ano o mes, usa esta pista si ayuda: {month_hint or "sin pista"}.
- Conserva el numero tal como aparece si hay duda, pero normalizalo a decimal cuando sea claro.
- No inventes valores. Si no puedes leer un numero, deja numero en blanco y baja la confianza.
- Usa formato ISO YYYY-MM-DD para fecha_iso cuando puedas deducir la fecha completa.

Formato exacto:
{{
  "rows": [
    {{
      "fecha": "texto visible de la fecha",
      "fecha_iso": "YYYY-MM-DD o vacio",
      "numero": "numero leido o vacio",
      "confianza": 0.0,
      "notas": "breve nota si hay duda"
    }}
  ]
}}
""".strip()

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": prompt},
                    image_content(image),
                ],
            }
        ],
    )

    data = extract_json(response.output_text)
    return data.get("rows", [])


def read_full_table_with_openai(
    client: OpenAI,
    image: Image.Image,
    model: str,
    month_hint: str,
) -> dict:
    prompt = f"""
Eres un extractor de tablas desde fotografias JPEG.
Lee esta foto y reconstruye la tabla con las mismas columnas visibles que tiene la foto.

Instrucciones:
- Devuelve solo JSON valido, sin markdown.
- Detecta los encabezados o nombres de columnas visibles. Si no hay encabezados claros, crea nombres descriptivos como "Columna 1", "Columna 2".
- Conserva absolutamente todas las columnas visibles, de izquierda a derecha.
- Extrae todos los numeros manuscritos de cada fila y colocalos en su columna correspondiente.
- No te limites a la fecha o a la primera columna numerica: cada interseccion fila-columna con un numero visible debe estar copiada.
- Extrae una fila por cada fila visible con datos y completa todas las columnas de esa fila.
- Devuelve cada fila como una lista ordenada de celdas, no como objeto/diccionario. Esto es importante si hay columnas con el mismo encabezado.
- Cada lista de fila debe tener exactamente la misma cantidad de elementos que "columns". Si una celda esta vacia o no se puede leer, pon "" en esa posicion.
- No inventes valores.
- Si una fecha no incluye ano o mes, usa esta pista si ayuda: {month_hint or "sin pista"}.
- Marca como dudosa cualquier celda borrosa, ambigua, parcialmente tapada, vacia por ilegible o con escritura dificil de leer.
- En doubtful_cells, incluye "column_index" usando indice 1-based de izquierda a derecha, ademas del nombre de columna.
- Usa texto tal como se ve, salvo normalizaciones obvias de numeros o fechas.

Formato exacto:
{{
  "columns": ["Nombre columna 1", "Nombre columna 2"],
  "rows": [
    ["valor columna 1", "valor columna 2"]
  ],
  "doubtful_cells": [
    {{"row": 1, "column": "Nombre columna 2", "column_index": 2, "reason": "motivo breve"}}
  ]
}}
""".strip()

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": prompt},
                    image_content(image),
                ],
            }
        ],
    )

    data = extract_json(response.output_text)
    return normalize_full_table_data(data)


def crop_wide_table_chunk(
    image: Image.Image,
    start_column: int,
    end_column: int,
    expected_columns: int,
    date_width_percent: int,
) -> Image.Image:
    width, height = image.size
    date_width = max(1, int(width * date_width_percent / 100))
    data_width = max(1, width - date_width)
    data_columns = max(1, expected_columns - 1)
    left = date_width + int((start_column - 2) / data_columns * data_width)
    right = date_width + int((end_column - 1) / data_columns * data_width)
    overlap = max(8, int(width * 0.01))
    left = max(date_width, left - overlap)
    right = min(width, right + overlap)

    date_crop = image.crop((0, 0, date_width, height))
    data_crop = image.crop((left, 0, right, height))
    combined = Image.new("RGB", (date_crop.width + data_crop.width, height), "white")
    combined.paste(date_crop, (0, 0))
    combined.paste(data_crop, (date_crop.width, 0))
    return ImageOps.expand(combined, border=20, fill="white")


def read_table_chunk_with_openai(
    client: OpenAI,
    image: Image.Image,
    model: str,
    month_hint: str,
    start_column: int,
    end_column: int,
) -> dict:
    chunk_columns = ["Fecha"] + [f"Columna {column}" for column in range(start_column, end_column + 1)]
    prompt = f"""
Eres un extractor de tablas desde una fotografia JPEG recortada por columnas.
La primera columna visible es la FECHA de referencia. A su derecha aparecen las columnas originales {start_column} a {end_column}.

Instrucciones:
- Devuelve solo JSON valido, sin markdown.
- Debes leer todos los numeros manuscritos de las columnas originales {start_column} a {end_column}.
- No te limites a la primera columna numerica.
- Devuelve exactamente estas columnas y en este orden: {chunk_columns}.
- Cada fila debe ser una lista con {len(chunk_columns)} elementos: fecha y luego los valores de cada columna.
- Si una celda esta vacia o no se puede leer, pon "" en esa posicion.
- Si una fecha no incluye ano o mes, usa esta pista si ayuda: {month_hint or "sin pista"}.
- Marca como dudosa cualquier celda borrosa, ambigua o dificil de leer.
- En doubtful_cells, usa column_index 1 para Fecha, 2 para Columna {start_column}, 3 para la siguiente, etc.

Formato exacto:
{{
  "columns": {chunk_columns},
  "rows": [
    ["fecha", "valor columna {start_column}", "valor columna siguiente"]
  ],
  "doubtful_cells": [
    {{"row": 1, "column": "Columna {start_column}", "column_index": 2, "reason": "motivo breve"}}
  ]
}}
""".strip()

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": prompt},
                    image_content(image),
                ],
            }
        ],
    )

    return normalize_full_table_data(extract_json(response.output_text))


def read_wide_table_with_openai(
    client: OpenAI,
    image: Image.Image,
    model: str,
    month_hint: str,
    expected_columns: int,
    date_width_percent: int,
) -> dict:
    all_columns = ["Fecha"] + [f"Columna {column}" for column in range(2, expected_columns + 1)]
    merged_rows: list[dict] = []
    merged_doubtful_cells: list[dict] = []

    for start_column in range(2, expected_columns + 1, DIRECT_CHUNK_SIZE):
        end_column = min(expected_columns, start_column + DIRECT_CHUNK_SIZE - 1)
        chunk_image = crop_wide_table_chunk(
            image=image,
            start_column=start_column,
            end_column=end_column,
            expected_columns=expected_columns,
            date_width_percent=date_width_percent,
        )
        chunk_data = read_table_chunk_with_openai(
            client=client,
            image=chunk_image,
            model=model,
            month_hint=month_hint,
            start_column=start_column,
            end_column=end_column,
        )

        for row_index, chunk_row in enumerate(chunk_data["rows"]):
            while len(merged_rows) <= row_index:
                merged_rows.append({column: "" for column in all_columns})
            merged_rows[row_index]["Fecha"] = merged_rows[row_index].get("Fecha") or chunk_row.get("Fecha", "")
            for column in range(start_column, end_column + 1):
                column_name = f"Columna {column}"
                merged_rows[row_index][column_name] = chunk_row.get(column_name, "")

        for doubtful_cell in chunk_data["doubtful_cells"]:
            chunk_column_index = pd.to_numeric(doubtful_cell.get("column_index", ""), errors="coerce")
            if pd.isna(chunk_column_index):
                continue
            chunk_column_index = int(chunk_column_index)
            if chunk_column_index == 1:
                global_column = "Fecha"
            else:
                global_column_number = start_column + chunk_column_index - 2
                global_column = f"Columna {global_column_number}"
            merged_doubtful_cells.append(
                {
                    **doubtful_cell,
                    "column": global_column,
                    "column_index": all_columns.index(global_column) + 1,
                }
            )

    return {
        "columns": all_columns,
        "rows": merged_rows,
        "doubtful_cells": merged_doubtful_cells,
    }


def make_unique_columns(columns: list[object]) -> list[str]:
    counts = {}
    unique_columns = []
    for index, column in enumerate(columns, start=1):
        base_name = str(column or "").strip() or f"Columna {index}"
        count = counts.get(base_name, 0) + 1
        counts[base_name] = count
        unique_columns.append(base_name if count == 1 else f"{base_name} ({count})")
    return unique_columns


def normalize_full_table_data(data: dict) -> dict:
    raw_columns = data.get("columns") or []
    raw_rows = data.get("rows") or []
    doubtful_cells = data.get("doubtful_cells") or []

    max_list_width = max(
        [len(row) for row in raw_rows if isinstance(row, list)] or [0]
    )
    while len(raw_columns) < max_list_width:
        raw_columns.append(f"Columna {len(raw_columns) + 1}")

    columns = make_unique_columns(raw_columns)
    normalized_rows = []

    for row in raw_rows:
        if isinstance(row, list):
            values = list(row)
            while len(values) < len(columns):
                values.append("")
            normalized_rows.append({columns[index]: values[index] for index in range(len(columns))})
        elif isinstance(row, dict):
            extra_columns = [column for column in row.keys() if str(column) not in raw_columns]
            if extra_columns:
                raw_columns.extend(extra_columns)
                columns = make_unique_columns(raw_columns)
                for existing_row in normalized_rows:
                    for column in columns:
                        existing_row.setdefault(column, "")
            normalized_rows.append(
                {
                    unique_column: row.get(raw_column, "")
                    for raw_column, unique_column in zip(raw_columns, columns)
                }
            )

    normalized_rows = [
        {column: row.get(column, "") for column in columns}
        for row in normalized_rows
    ]

    return {
        "columns": columns,
        "rows": normalized_rows,
        "doubtful_cells": normalize_doubtful_cells(doubtful_cells, columns),
    }


def normalize_doubtful_cells(doubtful_cells: list[dict], columns: list[str]) -> list[dict]:
    normalized = []
    for cell in doubtful_cells:
        column = cell.get("column", "")
        column_name = str(column)
        column_index = pd.to_numeric(cell.get("column_index", ""), errors="coerce")

        if not pd.isna(column_index):
            index = int(column_index) - 1
            if 0 <= index < len(columns):
                column_name = columns[index]

        normalized.append({**cell, "column": column_name})
    return normalized


def build_excel(df: pd.DataFrame, doubtful_cells: Optional[list[dict]] = None) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Datos")
        worksheet = writer.sheets["Datos"]
        red_font = Font(color="FF0000")
        column_positions = {str(column): index + 1 for index, column in enumerate(df.columns)}
        for doubtful_cell in doubtful_cells or []:
            row_number = int(doubtful_cell.get("row", 0) or 0)
            column_name = str(doubtful_cell.get("column", ""))
            column_number = column_positions.get(column_name)
            if row_number > 0 and column_number:
                worksheet.cell(row=row_number + 1, column=column_number).font = red_font
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 45)
    return output.getvalue()


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def find_header_map(worksheet) -> tuple[Optional[int], dict[str, int]]:
    aliases = {
        "fecha": {"fecha", "dia", "date"},
        "fecha_iso": {"fechaiso", "fechaiso", "fechaformato", "fechaexcel"},
        "numero": {"numero", "num", "valor", "cantidad", "importe", "n"},
        "confianza": {"confianza", "fiabilidad", "confidence"},
        "notas": {"notas", "nota", "observaciones", "observacion", "dudas"},
    }

    for row in worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row)):
        found = {}
        for cell in row:
            normalized = normalize_header(cell.value)
            for field, names in aliases.items():
                if normalized in names and field not in found:
                    found[field] = cell.column
        if "fecha" in found and "numero" in found:
            return row[0].row, found

    return None, {}


def next_append_row(worksheet, min_row: int) -> int:
    last_used_row = min_row - 1
    for row_number in range(min_row, worksheet.max_row + 1):
        has_content = any(
            worksheet.cell(row_number, column).value not in (None, "")
            for column in range(1, worksheet.max_column + 1)
        )
        if has_content:
            last_used_row = row_number
    return last_used_row + 1


def is_doubtful(row: pd.Series) -> bool:
    confidence = pd.to_numeric(row.get("confianza", ""), errors="coerce")
    notes = str(row.get("notas", "") or "").strip()
    number = str(row.get("numero", "") or "").strip()

    if not number:
        return True
    if pd.isna(confidence):
        return bool(notes)
    return float(confidence) < DOUBTFUL_CONFIDENCE or bool(notes)


def red_font_like(font: Font) -> Font:
    new_font = copy(font)
    new_font.color = "FF0000"
    return new_font


def build_preformatted_excel(df: pd.DataFrame, template_bytes: bytes) -> bytes:
    workbook = load_workbook(io.BytesIO(template_bytes))
    worksheet = workbook.active
    header_row, header_map = find_header_map(worksheet)

    if header_map:
        column_map = {
            "fecha": header_map.get("fecha_iso") or header_map.get("fecha"),
            "numero": header_map.get("numero"),
            "confianza": header_map.get("confianza"),
            "notas": header_map.get("notas"),
        }
        start_row = next_append_row(worksheet, (header_row or 1) + 1)
    else:
        column_map = {"fecha": 1, "numero": 2, "confianza": 3, "notas": 4}
        start_row = next_append_row(worksheet, 1)

    for index, row in df.reset_index(drop=True).iterrows():
        excel_row = start_row + index
        date_value = row.get("fecha_iso") or row.get("fecha") or ""
        number_value = row.get("numero", "")
        doubtful = is_doubtful(row)

        values = {
            "fecha": date_value,
            "numero": number_value,
            "confianza": row.get("confianza", ""),
            "notas": row.get("notas", ""),
        }

        for field, value in values.items():
            column = column_map.get(field)
            if not column:
                continue
            cell = worksheet.cell(excel_row, column)
            cell.value = value
            if field == "numero" and doubtful:
                cell.font = red_font_like(cell.font)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def find_header_map_from_rows(rows: list[list[object]]) -> tuple[Optional[int], dict[str, int]]:
    aliases = {
        "fecha": {"fecha", "dia", "date"},
        "fecha_iso": {"fechaiso", "fechaiso", "fechaformato", "fechaexcel"},
        "numero": {"numero", "num", "valor", "cantidad", "importe", "n"},
        "confianza": {"confianza", "fiabilidad", "confidence"},
        "notas": {"notas", "nota", "observaciones", "observacion", "dudas"},
    }

    for row_number, row_values in enumerate(rows[:20], start=1):
        found = {}
        for column_number, value in enumerate(row_values, start=1):
            normalized = normalize_header(value)
            for field, names in aliases.items():
                if normalized in names and field not in found:
                    found[field] = column_number
        if "fecha" in found and "numero" in found:
            return row_number, found

    return None, {}


def next_append_xls_row(rows: list[list[object]], min_row: int) -> int:
    last_used_row = min_row - 1
    for row_number in range(min_row, len(rows) + 1):
        row_values = rows[row_number - 1]
        if any(value not in (None, "") for value in row_values):
            last_used_row = row_number
    return last_used_row + 1


def build_preformatted_xls(df: pd.DataFrame, template_bytes: bytes) -> bytes:
    read_workbook = xlrd.open_workbook(file_contents=template_bytes, formatting_info=True)
    read_sheet = read_workbook.sheet_by_index(0)
    rows = [
        [read_sheet.cell_value(row_index, column_index) for column_index in range(read_sheet.ncols)]
        for row_index in range(read_sheet.nrows)
    ]
    header_row, header_map = find_header_map_from_rows(rows)

    if header_map:
        column_map = {
            "fecha": header_map.get("fecha_iso") or header_map.get("fecha"),
            "numero": header_map.get("numero"),
            "confianza": header_map.get("confianza"),
            "notas": header_map.get("notas"),
        }
        start_row = next_append_xls_row(rows, (header_row or 1) + 1)
    else:
        column_map = {"fecha": 1, "numero": 2, "confianza": 3, "notas": 4}
        start_row = next_append_xls_row(rows, 1)

    write_workbook = copy_xls_workbook(read_workbook)
    write_sheet = write_workbook.get_sheet(0)
    red_style = xlwt.easyxf("font: colour red;")

    for index, row in df.reset_index(drop=True).iterrows():
        excel_row = start_row + index
        date_value = row.get("fecha_iso") or row.get("fecha") or ""
        number_value = row.get("numero", "")
        doubtful = is_doubtful(row)

        values = {
            "fecha": date_value,
            "numero": number_value,
            "confianza": row.get("confianza", ""),
            "notas": row.get("notas", ""),
        }

        for field, value in values.items():
            column = column_map.get(field)
            if not column:
                continue
            if field == "numero" and doubtful:
                write_sheet.write(excel_row - 1, column - 1, value, red_style)
            else:
                write_sheet.write(excel_row - 1, column - 1, value)

    output = io.BytesIO()
    write_workbook.save(output)
    return output.getvalue()


def build_completed_workbook(df: pd.DataFrame, template_name: str, template_bytes: bytes) -> tuple[bytes, str, str]:
    lower_name = template_name.lower()
    if lower_name.endswith(".xls"):
        return (
            build_preformatted_xls(df, template_bytes),
            "fotos_ute_bilbao.xls",
            "application/vnd.ms-excel",
        )

    return (
        build_preformatted_excel(df, template_bytes),
        "fotos_ute_bilbao.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def app() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="UTE", layout="wide")
    st.title(APP_TITLE)

    with st.sidebar:
        st.header("Configuracion")
        api_key = st.text_input(
            "OPENAI_API_KEY",
            value=os.getenv("OPENAI_API_KEY", ""),
            type="password",
            help="Tambien puedes guardarla en un archivo .env.",
        )
        model = st.text_input("Modelo", value=DEFAULT_MODEL)
        selected_month = st.date_input(
            "Mes de referencia",
            value=date.today().replace(day=1),
            help="Ayuda si las fechas de la foto no muestran ano o mes.",
        )
        month_hint = selected_month.strftime("%B %Y")

    if not api_key:
        st.warning("Introduce una clave API de OpenAI para poder leer la foto.")

    preformatted_tab, direct_tab = st.tabs(["Excel preformado", "Foto a Excel"])

    with preformatted_tab:
        uploaded_file = st.file_uploader("Sube la foto JPEG", type=["jpg", "jpeg"], key="template_photo")
        uploaded_excel = st.file_uploader(
            "Sube el Excel preformado",
            type=["xlsx", "xls"],
            key="template_excel",
        )

        if not uploaded_file or not uploaded_excel:
            st.info("Sube una foto JPEG y un Excel preformado para generar la descarga final.")
        else:
            try:
                image = Image.open(uploaded_file).convert("RGB")
            except Exception as exc:
                st.error(f"No se pudo abrir la imagen JPEG: {exc}")
                return

            with st.expander("Vista previa de la foto", expanded=True):
                st.image(image, caption=uploaded_file.name, use_container_width=True)

            if st.button("Convertir a tabla", type="primary", key="template_convert"):
                if not api_key:
                    st.warning("Introduce una clave API de OpenAI para poder leer la foto.")
                    return

                client = OpenAI(api_key=api_key)
                progress = st.progress(0)
                status = st.empty()

                try:
                    status.write("Leyendo foto...")
                    all_rows = read_image_with_openai(
                        client=client,
                        image=image,
                        model=model,
                        month_hint=month_hint,
                    )
                    progress.progress(1.0)

                    if not all_rows:
                        st.warning("No se encontraron filas con fecha y numero.")
                        return

                    df = pd.DataFrame(all_rows)
                    preferred_columns = ["fecha", "fecha_iso", "numero", "confianza", "notas"]
                    df = df[[column for column in preferred_columns if column in df.columns]]
                    st.session_state["extracted_df"] = df
                    status.write("Lectura completada. Revisa la tabla antes de descargar.")
                except Exception as exc:
                    st.error(f"No se pudo convertir la foto: {exc}")

            if "extracted_df" in st.session_state:
                st.subheader("Revision")
                edited_df = st.data_editor(
                    st.session_state["extracted_df"],
                    num_rows="dynamic",
                    use_container_width=True,
                    key="edited_df",
                )
                excel_bytes, excel_name, excel_mime = build_completed_workbook(
                    edited_df,
                    uploaded_excel.name,
                    uploaded_excel.getvalue(),
                )
                st.download_button(
                    f"Descargar Excel completado ({excel_name.split('.')[-1]})",
                    data=excel_bytes,
                    file_name=excel_name,
                    mime=excel_mime,
                    on_click="ignore",
                    key="template_download",
                )

    with direct_tab:
        direct_file = st.file_uploader("Sube la foto JPEG", type=["jpg", "jpeg"], key="direct_photo")
        use_manual_columns = st.checkbox(
            "Indicar numero de columnas manualmente",
            value=False,
            help="Activalo si una foto ancha no detecta todas las columnas.",
        )
        expected_direct_columns = None
        date_width_percent = 18
        if use_manual_columns:
            expected_direct_columns = st.number_input(
                "Numero total de columnas en la foto",
                min_value=2,
                max_value=60,
                value=23,
                step=1,
                help="Incluye la columna de fecha. Por ejemplo: fecha + 22 columnas de numeros = 23.",
            )
            date_width_percent = st.slider(
                "Ancho aproximado de la columna Fecha",
                min_value=8,
                max_value=35,
                value=18,
                step=1,
                help="Ajustalo si la columna de fecha ocupa mas o menos espacio en la foto.",
            )

        if not direct_file:
            st.info("Sube una foto JPEG para convertirla directamente a Excel.")
            return

        try:
            direct_image = Image.open(direct_file).convert("RGB")
        except Exception as exc:
            st.error(f"No se pudo abrir la imagen JPEG: {exc}")
            return

        with st.expander("Vista previa de la foto", expanded=True):
            st.image(direct_image, caption=direct_file.name, use_container_width=True)

        if st.button("Convertir foto a Excel", type="primary", key="direct_convert"):
            if not api_key:
                st.warning("Introduce una clave API de OpenAI para poder leer la foto.")
                return

            client = OpenAI(api_key=api_key)
            progress = st.progress(0)
            status = st.empty()

            try:
                if use_manual_columns and expected_direct_columns and expected_direct_columns >= 12:
                    status.write(
                        f"Leyendo tabla ancha por tramos: {expected_direct_columns} columnas..."
                    )
                    extracted_table = read_wide_table_with_openai(
                        client=client,
                        image=direct_image,
                        model=model,
                        month_hint=month_hint,
                        expected_columns=int(expected_direct_columns),
                        date_width_percent=int(date_width_percent),
                    )
                else:
                    status.write("Leyendo columnas y datos de la foto...")
                    extracted_table = read_full_table_with_openai(
                        client=client,
                        image=direct_image,
                        model=model,
                        month_hint=month_hint,
                    )
                progress.progress(1.0)

                columns = extracted_table["columns"]
                rows = extracted_table["rows"]
                if not columns or not rows:
                    st.warning("No se encontraron datos tabulares en la foto.")
                    return

                st.session_state["direct_columns"] = columns
                st.session_state["direct_df"] = pd.DataFrame(rows, columns=columns)
                st.session_state["direct_doubtful_cells"] = extracted_table["doubtful_cells"]
                status.write(f"Lectura completada: {len(columns)} columnas detectadas. Revisa la tabla antes de descargar.")
            except Exception as exc:
                st.error(f"No se pudo convertir la foto: {exc}")

        if "direct_df" in st.session_state:
            st.subheader("Revision")
            st.caption(f"Columnas detectadas: {len(st.session_state['direct_df'].columns)}")
            edited_direct_df = st.data_editor(
                st.session_state["direct_df"],
                num_rows="dynamic",
                use_container_width=True,
                key="direct_edited_df",
            )
            direct_excel_bytes = build_excel(
                edited_direct_df,
                st.session_state.get("direct_doubtful_cells", []),
            )
            st.download_button(
                "Descargar Excel de la foto (.xlsx)",
                data=direct_excel_bytes,
                file_name="foto_ute_bilbao.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                on_click="ignore",
                key="direct_download",
            )


if __name__ == "__main__":
    app()
